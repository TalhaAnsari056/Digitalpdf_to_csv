from pathlib import Path

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)
from openpyxl.utils import get_column_letter


class ExcelFormatterService:

    ###############################################################
    # PUBLIC
    ###############################################################

    @classmethod
    def export(
        cls,
        dataframe: pd.DataFrame,
        output_file: Path,
    ):

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Balance Sheet"

        ###########################################################
        # Write dataframe
        ###########################################################

        cls._write_header(
            worksheet,
            dataframe.columns.tolist(),
        )

        cls._write_rows(
            worksheet,
            dataframe,
        )

        ###########################################################
        # Formatting
        ###########################################################

        cls._style_header(worksheet)

        cls._style_rows(
            worksheet,
            dataframe,
        )

        cls._merge_section_cells(
            worksheet,
            dataframe,
        )

        cls._merge_subsection_cells(
            worksheet,
            dataframe,
        )

        cls._format_amount_column(
            worksheet,
            dataframe,
        )

        cls._auto_width(
            worksheet,
        )

        cls._freeze_header(
            worksheet,
        )

        cls._enable_filter(
            worksheet,
        )

        ###########################################################

        workbook.save(output_file)

    ###############################################################
    # WRITE HEADER
    ###############################################################

    @staticmethod
    def _write_header(
        worksheet,
        columns,
    ):

        for col, name in enumerate(columns, start=1):

            worksheet.cell(
                row=1,
                column=col,
                value=name,
            )

    ###############################################################
    # WRITE DATA
    ###############################################################

    @staticmethod
    def _write_rows(
        worksheet,
        dataframe,
    ):

        for row in dataframe.itertuples(index=False):

            worksheet.append(list(row))

    ###############################################################
    # HEADER STYLE
    ###############################################################

    @staticmethod
    def _style_header(
        worksheet,
    ):

        fill = PatternFill(
            fill_type="solid",
            fgColor="404040",
        )

        font = Font(
            bold=True,
            color="FFFFFF",
        )

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        for cell in worksheet[1]:

            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = alignment

    ###############################################################
    # ROW STYLE
    ###############################################################

    @classmethod
    def _style_rows(
        cls,
        worksheet,
        dataframe,
    ):

        thin = Side(style="thin")

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

        section_fill = PatternFill(
            fill_type="solid",
            fgColor="D9D9D9",
        )

        subsection_fill = PatternFill(
            fill_type="solid",
            fgColor="EFEFEF",
        )

        total_fill = PatternFill(
            fill_type="solid",
            fgColor="FFF2CC",
        )

        for excel_row, (_, row) in enumerate(
            dataframe.iterrows(),
            start=2,
        ):

            row_type = str(row.get("row_type", "")).upper()

            for cell in worksheet[excel_row]:

                cell.border = border

                cell.alignment = Alignment(
                    vertical="center",
                )

            #######################################################

            if row_type == "SECTION":

                for cell in worksheet[excel_row]:

                    cell.font = Font(
                        bold=True,
                        size=12,
                    )

                    cell.fill = section_fill

            #######################################################

            elif row_type == "SUBSECTION":

                for cell in worksheet[excel_row]:

                    cell.font = Font(
                        bold=True,
                    )

                    cell.fill = subsection_fill

            #######################################################

            elif row_type == "TOTAL":

                for cell in worksheet[excel_row]:

                    cell.font = Font(
                        bold=True,
                    )

                    cell.fill = total_fill

    ###############################################################
    # MERGE SECTION
    ###############################################################

    @staticmethod
    def _merge_section_cells(
        worksheet,
        dataframe,
    ):

        start = 2

        current = None

        for idx, value in enumerate(
            dataframe["section"],
            start=2,
        ):

            if value != current:

                if current is not None:

                    worksheet.merge_cells(
                        start_row=start,
                        end_row=idx - 1,
                        start_column=2,
                        end_column=2,
                    )

                current = value
                start = idx

        worksheet.merge_cells(
            start_row=start,
            end_row=len(dataframe) + 1,
            start_column=2,
            end_column=2,
        )

    ###############################################################
    # MERGE SUBSECTION
    ###############################################################

    @staticmethod
    def _merge_subsection_cells(
        worksheet,
        dataframe,
    ):

        start = None

        previous_section = None

        previous_subsection = None

        for excel_row, (_, row) in enumerate(
            dataframe.iterrows(),
            start=2,
        ):

            section = row["section"]

            subsection = row["subsection"]

            if subsection == "":

                continue

            if subsection != previous_subsection or section != previous_section:

                if start is not None and excel_row - start > 1:

                    worksheet.merge_cells(
                        start_row=start,
                        end_row=excel_row - 1,
                        start_column=3,
                        end_column=3,
                    )

                start = excel_row

                previous_section = section

                previous_subsection = subsection

        if start is not None:

            worksheet.merge_cells(
                start_row=start,
                end_row=len(dataframe) + 1,
                start_column=3,
                end_column=3,
            )

    ###############################################################
    # AMOUNT FORMAT
    ###############################################################

    @staticmethod
    def _format_amount_column(
        worksheet,
        dataframe,
    ):

        if "amount" not in dataframe.columns:
            return

        amount_col = dataframe.columns.get_loc("amount") + 1

        for row in range(
            2,
            len(dataframe) + 2,
        ):

            cell = worksheet.cell(
                row=row,
                column=amount_col,
            )

            value = str(cell.value).strip()

            if value == "":
                continue

            try:

                number = float(value)

                cell.value = number

                cell.number_format = "#,##0.00"

            except Exception:

                pass

    ###############################################################
    # AUTO WIDTH
    ###############################################################

    @staticmethod
    def _auto_width(
        worksheet,
    ):

        for column in worksheet.columns:

            maximum = 0

            letter = get_column_letter(
                column[0].column,
            )

            for cell in column:

                try:

                    maximum = max(
                        maximum,
                        len(str(cell.value)),
                    )

                except Exception:

                    pass

            worksheet.column_dimensions[letter].width = min(
                maximum + 4,
                50,
            )

    ###############################################################
    # FREEZE
    ###############################################################

    @staticmethod
    def _freeze_header(
        worksheet,
    ):

        worksheet.freeze_panes = "A2"

    ###############################################################
    # FILTER
    ###############################################################

    @staticmethod
    def _enable_filter(
        worksheet,
    ):

        worksheet.auto_filter.ref = worksheet.dimensions
