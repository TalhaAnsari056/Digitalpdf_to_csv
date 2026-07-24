from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


class BankStatementExcelFormatterService:

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF",
    )

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @classmethod
    def export(
        cls,
        dataframe: pd.DataFrame,
        output_file: Path,
    ):

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Bank Statement"

        ###########################################################
        # Header
        ###########################################################

        headers = list(dataframe.columns)

        for column_index, header in enumerate(headers, start=1):

            cell = sheet.cell(
                row=1,
                column=column_index,
                value=header.replace("_", " ").title(),
            )

            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.border = cls.THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        ###########################################################
        # Data
        ###########################################################

        numeric_columns = {
            "debit",
            "credit",
            "balance",
        }

        for row in dataframe.itertuples(index=False):

            sheet.append(list(row))

        ###########################################################
        # Cell Formatting
        ###########################################################

        for row in sheet.iter_rows(min_row=2):

            for cell in row:

                cell.border = cls.THIN_BORDER

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

                header = headers[cell.column - 1]

                ###################################################
                # Numeric formatting
                ###################################################

                if header in numeric_columns:

                    value = cell.value

                    if value not in (
                        "",
                        None,
                    ):

                        try:

                            numeric = float(str(value).replace(",", ""))

                            cell.value = numeric

                            cell.number_format = "#,##0.00"

                            cell.alignment = Alignment(
                                horizontal="right",
                                vertical="top",
                            )

                        except Exception:
                            pass

        ###########################################################
        # Auto width
        ###########################################################

        for column_cells in sheet.columns:

            max_length = 0

            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:

                try:

                    value = "" if cell.value is None else str(cell.value)

                    if len(value) > max_length:
                        max_length = len(value)

                except Exception:
                    pass

            width = min(
                max(max_length + 3, 12),
                60,
            )

            sheet.column_dimensions[column_letter].width = width

        ###########################################################
        # Better widths
        ###########################################################

        preferred_widths = {
            "row_type": 22,
            "date": 18,
            "description": 55,
            "debit": 18,
            "credit": 18,
            "balance": 18,
            "currency": 12,
        }

        for index, header in enumerate(headers, start=1):

            if header in preferred_widths:

                sheet.column_dimensions[get_column_letter(index)].width = (
                    preferred_widths[header]
                )

        ###########################################################
        # Freeze
        ###########################################################

        sheet.freeze_panes = "A2"

        ###########################################################
        # Filter
        ###########################################################

        sheet.auto_filter.ref = sheet.dimensions

        ###########################################################
        # Row Height
        ###########################################################

        for row in range(2, sheet.max_row + 1):

            sheet.row_dimensions[row].height = 22

        ###########################################################
        # Highlight Row Types
        ###########################################################

        row_type_index = None

        for i, header in enumerate(headers):

            if header == "row_type":

                row_type_index = i + 1

                break

        if row_type_index:

            opening_fill = PatternFill(
                "solid",
                fgColor="D9EAD3",
            )

            closing_fill = PatternFill(
                "solid",
                fgColor="F4CCCC",
            )

            summary_fill = PatternFill(
                "solid",
                fgColor="FFF2CC",
            )

            for row in range(2, sheet.max_row + 1):

                value = str(
                    sheet.cell(
                        row=row,
                        column=row_type_index,
                    ).value
                ).upper()

                if value == "OPENING_BALANCE":

                    fill = opening_fill

                elif value == "CLOSING_BALANCE":

                    fill = closing_fill

                elif value == "SUMMARY":

                    fill = summary_fill

                else:

                    continue

                for cell in sheet[row]:

                    cell.fill = fill

        ###########################################################
        # Save
        ###########################################################

        workbook.save(output_file)

        return output_file
